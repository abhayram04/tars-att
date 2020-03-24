
from openpyxl import Workbook, load_workbook
from cryptography.fernet import Fernet
from mkxl import sadd
import getpass
import os
def regie():

    name = input("Enter username: ")
    pwd = getpass.getpass(prompt="Password: ")

    file = open('key.txt','rb')
    key = file.read()
    if(not key):
        key = Fernet.generate_key()
        file = open('key.txt','wb')
        file.write(key)
        file.close()
        key = file.read()

    file.close()

    mp = os.path.abspath('../')
    path = mp+"/sub/"+name+"/"
    os.mkdir(path)
    f = Fernet(key)
    uenc = f.encrypt(name.encode())
    penc = f.encrypt(pwd.encode())


    wb = load_workbook("cap.xlsx")

    page = wb.active

    m_row = page.max_row

    n_cell = page.cell(row=m_row+1,column=1)
    n_cell.value = uenc

    m_cell = page.cell(row=m_row+1,column=2)
    m_cell.value = penc

    wb.save("cap.xlsx")
    
    sadd(name)

    return 0

    

    

