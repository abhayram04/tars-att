
import sys
import os
sys.path.append(os.path.abspath('../'))
from openpyxl import Workbook, load_workbook
from cryptography.fernet import Fernet
from main import dev_manner

wb = load_workbook("cap.xlsx")
page = wb.active


file = open('key.txt','rb')
key = file.read()
file.close()

f = Fernet(key)

def reg(uad,pad):

    m_row = page.max_row
    #print(m_row)
    for i in range(1,m_row+1):
        val = page.cell(row=i,column=1)
        fval = val.value
        dec = f.decrypt(fval.encode())
        dam = dec.decode()
        if(dam==uad):
            pval = page.cell(row=i,column=2)
            pal = pval.value
            pdec = f.decrypt(pal.encode())
            pam = pdec.decode()
            if(pam==pad):
                print("\n")
                print("LOGIN SUCCESSFUL. Welcome to Attendance app!")
                dev_manner(dam)
    print("\nInvalid credentials. Please try again!")

