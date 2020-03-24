
#This file holds the code for appending/adding names of students to a list
#File must be addded to the main program.
#Copyright TARS Inc.
#Made with ❤ in Mary's 

from openpyxl import Workbook,load_workbook
from schoose import scho,sadd


def remo(bam,ans):
    sam=bam
    sub = sadd(sam,ans)
    print("\n")
    name = input("Enter the name of the student to remove: ")
    print("Class selected: "+ans)


    for i in sub:
        path = "../sub/"+sam+"/"+ans+"/"+i+".xlsx"
        wb = load_workbook(path)

        page = wb.active

        co = 0
        m_row = page.max_row
        for x in range(1,m_row+1):
            n_cell=page.cell(row=x,column=2)
            vax=n_cell.value
            if(name==vax):
                page.delete_rows(x,1)
                co = 1

    if(co==0):
        print("\nSorry, student not found.\n")

            

   


        wb.save(path)

    