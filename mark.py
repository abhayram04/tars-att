
#This file contains the code required for marking attendace.
#File must be addded to the main program.
#Copyright TARS Inc.
#Made with ❤ in Mary's 


from openpyxl import Workbook, load_workbook
from datetime import date
from schoose import scho

today = date.today()
t_date = today.strftime("%d/%m/%y")

def marker(bam,ans):
    sam=bam
    print("\n")
    
    print("Class selected: "+ans)
    
    choc = scho(sam,ans)

    path = "../sub/"+sam+"/"+ans+"/"+choc+".xlsx"
    wb = load_workbook(path)

    #print("Subjects: 1. Algebra 2. Python 3.Machine Learning 4.Hindi 5.Spanish")
    #s_inp = input("Enter subject option: ")

    #The conditional statements are used to load *only* the file pertaining to the selected subject file.
    """if(s_inp=='1'):
        wb = load_workbook("../sub/"+sam+"/"+ans+"/math.xlsx")
    elif(s_inp=='2'):
        wb = load_workbook("../sub/"+sam+"/py.xlsx")
    elif(s_inp=='3'):
        wb = load_workbook("../sub/"+sam+"/ml.xlsx")
    elif(s_inp=='4'):
        wb = load_workbook("../sub/"+sam+"/hind.xlsx")
    elif(s_inp=='5'):
        wb = load_workbook("../sub/"+sam+"/span.xlsx")
    else:
        print("Wrong input")
        exit()"""
    
    page = wb.active
    m_row = page.max_row
    m_col = page.max_column

    #The next two lines of code set the date for the day in the excel sheet. 
    date_setter = page.cell(row=1,column=m_col+1)
    date_setter.value = t_date

    #Loops used to mark attendance for all the students listed in the excel sheet.
    for i in range(2,m_row+1):

        c_val = page.cell(row=i,column=2)
        cox=1
        while(cox==1):
            att = input("Enter attendance (P or A) for "+c_val.value+" :")
            att=att.upper()
            if(att=='P'):
                nik=0
                n_cell = page.cell(row=i,column=m_col+1)
                n_cell.value = att
            elif(att=='A'):
                nik=0
                n_cell = page.cell(row=i,column=m_col+1)
                n_cell.value = att
            else:
                nik=1
                print("Invalid input. Please try again.\n")
            cox=int(nik)
        print("\nATTENDANCE MARKED!!!\n")
        #n_cell = page.cell(row=i,column=m_col+1)
        #n_cell.value = att
        #Testing purpose
        #print(m_row)
        #print(m_col)

        #The conditional statements below are used to save *only* to the selected subject file.
        wb.save(path)
        """if(s_inp=='1'):
            wb.save("../sub/"+sam+"/math.xlsx")
        elif(s_inp=='2'):
            wb.save("../sub/"+sam+"/py.xlsx")
    
        elif(s_inp=='3'):
            wb.save("../sub/"+sam+"/ml.xlsx")
        
        elif(s_inp=='4'):
            wb.save("../sub/"+sam+"/hind.xlsx")
          
        elif(s_inp=='5'):
            wb.save("../sub/"+sam+"/span.xlsx")"""
        

        
        