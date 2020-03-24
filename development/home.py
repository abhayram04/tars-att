import getpass
from openpyxl import Workbook, load_workbook
from cryptography.fernet import Fernet
import os
from datetime import date

today = date.today()
t_date = today.strftime("%d/%m/%y")
wb = Workbook()

def bas_txt(path):
    wb=load_workbook(path)
    page = wb.active
    m_cell = page.cell(row=1,column=1)
    m_cell.value="RNo"
    m_cell = page.cell(row=1,column=2)
    m_cell.value="Name"
    wb.save(path)

def list_class(val,name):
    
    path = "./sub/"+name+"/"+"subs.txt"
    myFile = open(path,'w')
    for j in val:
        x=str(j)
        myFile.write(x)
        myFile.write('\n')
    myFile.close()

def list_sub(val,name,j):
    
    path = "./sub/"+name+"/"+j+"/subs.txt"
    myFile = open(path,'w')
    for j in val:
        x=str(j)
        myFile.write(x)
        myFile.write('\n')
    myFile.close()


def sadde(name):
    x=input("\nHow many classes do you teach for : ")
    x=int(x)
    clan = []
    
    print("\nEnter a class name followed by the Enter key and repeat:")
    for i in range(0,x):
        clin = input("")
        clan.append(clin)
        path = "./sub/"+name+"/"+clin
        list_class(clan,name)
        os.mkdir(path)

    
    for j in clan:
        s_list = []
        print("\n")

        subs = input("How many subjects do you teach for "+j+" class? : ")
        subs=int(subs)
        print("\nEnter a subject name (space & special characters not allowed) followed by the Enter key and repeat:")
        for nos in range(0,subs):
            
            inne = input("")
            path = "./sub/"+name+"/"+j+"/"+inne+".xlsx"
            wb.save(path)
            s_list.append(inne)
            
            bas_txt(path)
        list_sub(s_list,name,j)


def asker():
    x = input("Do you wish to continue using the app? (1-Yes, 0-No): ")
    return x
    

def adder(bam,ans):
    sam=bam
    #sub = ['math','span','ml','hind','py']
    sub = sadd(sam,ans)
    print("\n")
    name = input("Enter the name of the student: ")    

    for i in sub:
        path = "./sub/"+sam+"/"+ans+"/"+i+".xlsx"
        wb = load_workbook(path)

        page = wb.active

        m_row = page.max_row
        m_col = page.max_column
        
   
        #Function call to get max-row used in the excel sheet to be added in future release
        n_cell = page.cell(row=m_row+1,column=1)
        if(m_row==1):
            n_cell.value = m_row
        else:
            teval = page.cell(row=m_row,column=1)
            tev = teval.value
            n_cell.value = tev+1
        m_cell = page.cell(row=m_row+1,column=2)
        m_cell.value = name

        wb.save(path)

def marker(bam,ans):
    sam=bam
    print("\n")
    
    print("Class selected: "+ans)
    
    choc = scho(sam,ans)

    path = "./sub/"+sam+"/"+ans+"/"+choc+".xlsx"
    wb = load_workbook(path)

   
    #The conditional statements are used to load *only* the file pertaining to the selected subject file.

    
    page = wb.active
    m_row = page.max_row
    m_col = page.max_column

    #The next two lines of code set the date for the day in the excel sheet. 
    date_setter = page.cell(row=1,column=m_col+1)
    date_setter.value = t_date

    #Loops used to mark attendance for all the students listed in the excel sheet.
    for i in range(2,m_row+1):

        c_val = page.cell(row=i,column=2)
        cox=1
        while(cox==1):
            att = input("Enter attendance (P or A) for "+c_val.value+" :")
            att=att.upper()
            if(att=='P'):
                nik=0
                att="PRESENT"
                n_cell = page.cell(row=i,column=m_col+1)
                n_cell.value = att
            elif(att=='A'):
                nik=0
                att="ABSENT"
                n_cell = page.cell(row=i,column=m_col+1)
                n_cell.value = att
            else:
                nik=1
                print("Invalid input. Please try again.\n")
            cox=int(nik)
        print("\nATTENDANCE MARKED!!!\n")
        
        wb.save(path)
        
        
def scho(dam,nam):
    print("Your subjects-")
    path = "./sub/"+dam+"/"+nam+"/subs.txt"
    lines = tuple(open(path).read().split('\n'))
    ran = len(lines)
    for x in range(1,ran):
        y=str(x)
        x=x-1
        x=str(lines[x])
        print(y+". "+x)
        
    mans = input("Select subject (number): ")
    mans = int(mans)
    mans = mans-1
    answer = lines[mans]
    answer = str(answer)

    return answer

def sadd(dam,nam):
    path = "./sub/"+dam+"/"+nam+"/subs.txt"
    lines = tuple(open(path).read().split('\n'))
    ran = len(lines)
    act = []
    for x in range(1,ran):
        y=str(x)
        x=x-1
        x=str(lines[x])
        act.append(x)
    return act

def remo(bam,ans):
    sam=bam
    sub = sadd(sam,ans)
    print("\n")
    name = input("Enter the name of the student to remove: ")
    name = str(name)
    for i in sub:
        path = "./sub/"+sam+"/"+ans+"/"+i+".xlsx"
        wb = load_workbook(path)
        page = wb.active

        co = 0
        m_row = page.max_row
        for x in range(2,m_row+1):
            n_cell=page.cell(row=x,column=2)
            vax=n_cell.value
            vax = str(vax)
            if(name==vax):
                page.delete_rows(x,1)
                co = 1
        wb.save(path)

    if(co==0):
        print("\nSorry, student not found.\n")

    

def tview(bam,ans):
    sam = bam
    print("\n")
    #print("Subjects: 1.)Algebra 2.)Python 3.)Machine Learning 4.)Hindi 5.)Spanish")
    #s_inp = input("Enter subject option: ")
    print("Class selected: "+ans)
    
    choc = scho(sam,ans)

    path = "./sub/"+sam+"/"+ans+"/"+choc+".xlsx"
    wb = load_workbook(path)

    #The conditional statements are used to load *only* the file pertaining to the selected subject file.

    page = wb.active

    m_row = page.max_row
    m_col = page.max_column

    print("\n")
    for i in range(1,m_row+1):
        for j in range(1,m_col+1):
            c_val = page.cell(row=i,column=j)
            
            print(c_val.value,end=" |")
        print("\n")


def dev_manner(dam):
    
    xc = 0
    while(xc==0):
        print("\nYour classes-")
        path = "./sub/"+dam+"/"+"subs.txt"
        lines = tuple(open(path).read().split('\n'))
        ran = len(lines)
        
        for x in range(1,ran):
            y=str(x)
            x=x-1
            x=str(lines[x])
            print(y+". "+x)
        
        mans = input("Select class (number): ")
        mans = int(mans)

        if(mans>=1 and mans<ran):
            mans = mans-1
            answer = lines[mans]
            answer = str(answer)
            xc=1
        else:
            xc=0
            print("\nInvalid input. Please try again.\n")
        



    val=1
    while(val==1):
        os.system('cls')
        gam = dam
        print("\nClass selected: "+answer)
        print("\n")
        print("1.)Change class 2.)Add students 3.)Remove students 4.)Mark attendance 5.)View attendance 6.)LOGOUT")

        inp = input("Enter option (1-4):")

        if(inp=='1'):
            dev_manner(gam)
        if(inp=='2'):
            #Function call to add students to list
            adder(gam,answer)
            y=asker()

        elif(inp=='3'):
            #Function call to remove student from list
            remo(gam,answer)
            y=asker()

        elif(inp=='4'):
            #Function call to mark attendance of students
            marker(gam,answer)

            y=asker()

        elif(inp=='5'):
            #Function call to view attendance
            tview(gam,answer)
            y=asker()

        elif(inp=='6'):
            #Function call to logout from the service.
            os.system('cls')
            print("\n\nYou have successfully logged out.\n\n")
            homie()
            


        else:
            #Function call to print error message and restart program
            print("Incorrect choice. Please try again.")

            #Ticket - add the option to restart program again.
        y=int(y)
        if(y==0):
            os.system('cls')
            print("\n\nHope you enjoyed using the TARS app :D\n\n")
            homie()
        while(y>1):
            print("\nInvalid choice. Please try again.")
            y = asker()
            y=int(y)
        val=int(y)

        

def reg(stat):

    os.system('cls')

    if(stat==1):
        print("\nREGISTRATION COMPLETE!!! YOU CAN LOGIN NOW :)\n")

    print("\nLOGIN\n")
    uad = input("Username: ")
    pad = getpass.getpass("Password: ")
    wb = load_workbook("./sub/cap.xlsx")
    page = wb.active


    file = open('./sub/key.txt','rb')
    key = file.read()
    file.close()

    f = Fernet(key)
    m_row = page.max_row
    #print(m_row)
    for i in range(1,m_row+1):
        val = page.cell(row=i,column=1)
        fval = val.value
        dec = f.decrypt(fval.encode())
        dam = dec.decode()
        if(dam==uad):
            pval = page.cell(row=i,column=2)
            pal = pval.value
            pdec = f.decrypt(pal.encode())
            pam = pdec.decode()
            if(pam==pad):
                print("\n")
                print("LOGIN SUCCESSFUL. Welcome to Attendance app!")
                dev_manner(dam)
    print("\nInvalid credentials. Please try again!")
    homie()

def regie():
    os.system('cls')
    print("\nREGISTER\n")
    name = input("Enter username: ")
    pwd = getpass.getpass(prompt="Password: ")

    file = open('./sub/key.txt','rb')
    key = file.read()
    if(not key):
        key = Fernet.generate_key()
        file = open('./sub/key.txt','wb')
        file.write(key)
        file.close()
        key = file.read()

    file.close()

    path = "./sub/"+name+"/"
    os.mkdir(path)
    f = Fernet(key)
    uenc = f.encrypt(name.encode())
    penc = f.encrypt(pwd.encode())


    wb = load_workbook("./sub/cap.xlsx")

    page = wb.active

    m_row = page.max_row

    n_cell = page.cell(row=m_row+1,column=1)
    n_cell.value = uenc

    m_cell = page.cell(row=m_row+1,column=2)
    m_cell.value = penc

    wb.save("./sub/cap.xlsx")
    
    sadde(name)

    return 0

def homie():
    print("#############################################")
    print("##                 TARS                    ##")
    print("##            ATTENDANCE APP               ##")
    print("#############################################")
    print("Made with ❤ in Mary's")

    
    print("\n") 
    choi = input(" 1. LOGIN \n 2. REGISTER \n 3. README \n 4. ABOUT \n 5. EXIT \n\nEnter option: ")
    
    if(choi=='1'):
        
        reg(0)
    elif(choi=='2'):
        v = regie()
        if(v==0):
            reg(1)
    elif(choi=='3'):
        os.system('cls')
        path = 'readme.txt'
        lines = tuple(open(path,encoding="utf8").read().split('\n'))
        ran = len(lines)
        
        for x in range(1,ran):
            tee=str(lines[x])
            print(tee)
        tex = input("\nEnter 1 to return to main menu: ")
        tex = int(tex)
        if(tex==1):
            os.system('cls')
            homie()
        else:
            exit()
    elif(choi=='4'):
        os.system('cls')
        path = 'about.txt'
        lines = tuple(open(path,encoding="utf8").read().split('\n'))
        ran = len(lines)
        
        for x in range(1,ran):
            tee=str(lines[x])
            print(tee)
        tex = input("\nEnter 1 to return to main menu: ")
        tex = int(tex)
        if(tex==1):
            os.system('cls')
            homie()
        else:
            exit()
    elif(choi=='5'):
        exit()

    else:
        print("\nSorry, wrong input :/")
        print("Please try again!")
        homie()
        
homie()

    

       
        
    




