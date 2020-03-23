import os
from openpyxl import Workbook, load_workbook
wb = Workbook()

def bas_txt(path):
    wb=load_workbook(path)
    page = wb.active
    m_cell = page.cell(row=1,column=1)
    m_cell.value="RNo"
    m_cell = page.cell(row=1,column=2)
    m_cell.value="Name"
    wb.save(path)

def list_class(val,name):
    
    mp = os.path.abspath('../')
    path = mp+"/sub/"+name+"/"+"subs.txt"
    myFile = open(path,'w')
    for j in val:
        x=str(j)
        myFile.write(x)
        myFile.write('\n')
    myFile.close()

def list_sub(val,name,j):
    
    mp = os.path.abspath('../')
    path = mp+"/sub/"+name+"/"+j+"/subs.txt"
    myFile = open(path,'w')
    for j in val:
        x=str(j)
        myFile.write(x)
        myFile.write('\n')
    myFile.close()


def sadd(name):
    x=input("How many classes do you teach for?")
    x=int(x)
    clan = []
    
    print("Enter a class name (don't use space or special characters) followed by the Enter key and repeat:")
    for i in range(0,x):
        clin = input("")
        clan.append(clin)
        mp = os.path.abspath('../')
        path = mp+"/sub/"+name+"/"+clin
        list_class(clan,name)
        os.mkdir(path)
            

    

    print("\n")
    
    for j in clan:
        s_list = []
        print("How many subjects do you teach for "+j+"?")
        subs = input("")
        subs=int(subs)
        print("Enter a subject name (don't use space or special characters) followed by the Enter key and repeat:")
        for nos in range(0,subs):
            
            inne = input("")
            mp = os.path.abspath('../')
            path = mp+"/sub/"+name+"/"+j+"/"+inne+".xlsx"
            wb.save(path)
            s_list.append(inne)
            
            bas_txt(path)
        list_sub(s_list,name,j)

            

            

