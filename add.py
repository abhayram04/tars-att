
#This file holds the code for appending/adding names of students to a list
#File must be addded to the main program.
#Copyright TARS Inc.
#Made with ❤ in Mary's 

from openpyxl import Workbook,load_workbook


def adder():

    sub = ['math','span','ml','hind','py']
    print("\n")
    name = input("Enter the name of the student: ")

    for i in sub:
        path="../sub/"+i+".xlsx"
        wb = load_workbook(path)

        page = wb.active

        m_row = page.max_row
        m_col = page.max_column
        
   
        #Function call to get max-row used in the excel sheet to be added in future release
    
        n_cell = page.cell(row=m_row+1,column=1)
        n_cell.value = m_row+1

        m_cell = page.cell(row=m_row+1,column=2)
        m_cell.value = name

        wb.save(path)

    