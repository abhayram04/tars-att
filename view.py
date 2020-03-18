#This file contains the code necessary for viewing attendance of students.
#File must be addded to the main program.
#Copyright TARS Inc.
#Made with ❤ in Mary's 


from openpyxl import load_workbook

def tview(bam):
    sam = bam
    print("\n")
    print("Subjects: 1.)Algebra 2.)Python 3.)Machine Learning 4.)Hindi 5.)Spanish")
    s_inp = input("Enter subject option: ")

    #The conditional statements are used to load *only* the file pertaining to the selected subject file.
    if(s_inp=='1'):
        wb = load_workbook("../sub/"+sam+"/math.xlsx")
    elif(s_inp=='2'):
        wb = load_workbook("../sub/"+sam+"/py.xlsx")
    elif(s_inp=='3'):
        wb = load_workbook("../sub/"+sam+"/ml.xlsx")
    elif(s_inp=='4'):
        wb = load_workbook("../sub/"+sam+"/hind.xlsx")
    elif(s_inp=='5'):
        wb = load_workbook("../sub/"+sam+"/span.xlsx")
    else:
        print("Wrong input")

    page = wb.active

    m_row = page.max_row
    m_col = page.max_column

    print("\n")
    for i in range(1,m_row+1):
        for j in range(1,m_col+1):
            c_val = page.cell(row=i,column=j)
            
            print(c_val.value,end=" |")
        print("\n")